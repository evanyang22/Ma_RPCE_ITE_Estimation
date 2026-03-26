"""
Convert .npz file to Excel (.xlsx) format
Handles multiple arrays of different shapes by creating separate sheets
"""
 
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
 
def npz_to_xlsx(npz_path, xlsx_path):
    """
    Convert a .npz file to an Excel file with multiple sheets.
    
    Parameters:
    -----------
    npz_path : str
        Path to the input .npz file
    xlsx_path : str
        Path to the output .xlsx file
    """
    # Load the npz file
    data = np.load(npz_path)
    
    # Create Excel writer
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        
        # Process each array in the npz file
        for key in data.files:
            arr = data[key]
            
            print(f"Processing '{key}' - Shape: {arr.shape}, Dtype: {arr.dtype}")
            
            # Handle scalar values
            if arr.shape == () or arr.ndim == 0:
                df = pd.DataFrame({'value': [arr.item()]})
                df.to_excel(writer, sheet_name=key, index=False)
            
            # Handle 1D arrays
            elif arr.ndim == 1:
                df = pd.DataFrame({key: arr})
                df.to_excel(writer, sheet_name=key, index=False)
            
            # Handle 2D arrays
            elif arr.ndim == 2:
                # Create column names based on array shape
                if arr.shape[1] <= 100:  # If not too many columns
                    df = pd.DataFrame(arr, 
                                     columns=[f'col_{i}' for i in range(arr.shape[1])])
                else:
                    # For very wide arrays, just use default column names
                    df = pd.DataFrame(arr)
                df.to_excel(writer, sheet_name=key, index=True, index_label='row')
            
            # Handle 3D arrays (flatten to 2D)
            elif arr.ndim == 3:
                # Reshape 3D array: (dim0, dim1, dim2) -> create a 2D table
                # with dim0*dim2 rows and dim1 columns
                reshaped = arr.transpose(0, 2, 1).reshape(-1, arr.shape[1])
                
                # Create multi-index for better organization
                idx_level0 = np.repeat(range(arr.shape[0]), arr.shape[2])
                idx_level1 = np.tile(range(arr.shape[2]), arr.shape[0])
                
                df = pd.DataFrame(reshaped,
                                 columns=[f'feature_{i}' for i in range(arr.shape[1])],
                                 index=pd.MultiIndex.from_arrays([idx_level0, idx_level1],
                                                                  names=['sample', 'replicate']))
                df.to_excel(writer, sheet_name=key)
            
            # Handle higher dimensional arrays
            else:
                # Flatten to 2D
                reshaped = arr.reshape(arr.shape[0], -1)
                df = pd.DataFrame(reshaped)
                df.to_excel(writer, sheet_name=key, index=True)
                print(f"  Note: {arr.ndim}D array flattened to 2D")
    
    print(f"\nSuccessfully created: {xlsx_path}")
    print(f"Total sheets: {len(data.files)}")
 

 
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
 
 
def tensors_to_excel(*tensor_lists, list_names=None, output_path='tensors_output.xlsx'):
    """
    Convert multiple lists of tensors into an Excel file with one sheet per tensor index.
    Each sheet contains the i-th tensor from all lists as columns.
    
    Args:
        *tensor_lists: Variable number of lists of tensors (all must have same length)
        list_names: List of strings for column headers (optional, defaults to "Column_1", "Column_2", etc.)
        output_path: Path to save the Excel file
    
    Returns:
        str: Path to the saved Excel file
        
    Example:
        >>> import torch
        >>> list1 = [torch.randn(10) for _ in range(100)]
        >>> list2 = [torch.randn(10) for _ in range(100)]
        >>> list3 = [torch.randn(10) for _ in range(100)]
        >>> list4 = [torch.randn(10) for _ in range(100)]
        >>> list5 = [torch.randn(10) for _ in range(100)]
        >>> list6 = [torch.randn(10) for _ in range(100)]
        >>> 
        >>> names = ['OBS_Original', 'RCT_Original', 'OBS_Projected', 
        ...          'RCT_Projected', 'Distance_Before', 'Distance_After']
        >>> 
        >>> tensors_to_excel(list1, list2, list3, list4, list5, list6,
        ...                  list_names=names, output_path='output.xlsx')
    """
    if len(tensor_lists) == 0:
        raise ValueError("At least one tensor list must be provided")
    
    num_lists = len(tensor_lists)
    
    # Generate default names if not provided
    if list_names is None:
        list_names = [f'Column_{i+1}' for i in range(num_lists)]
    
    # Validation
    if len(list_names) != num_lists:
        raise ValueError(f"list_names must contain exactly {num_lists} names (got {len(list_names)})")
    
    num_tensors = len(tensor_lists[0])
    for i, lst in enumerate(tensor_lists):
        if len(lst) != num_tensors:
            raise ValueError(f"List {i+1} has {len(lst)} tensors, expected {num_tensors}")
    
    wb = Workbook()
    wb.remove(wb.active)
    
    for i in range(num_tensors):
        sheet = wb.create_sheet(f'Tensor_{i+1}')
        
        # Write headers with formatting
        for col_idx, name in enumerate(list_names, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=name)
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center')
        
        tensors_at_i = [lst[i] for lst in tensor_lists]
        
        # Write tensor data
        for col_idx, tensor in enumerate(tensors_at_i, start=1):
            if tensor.dim() == 0:  # Scalar
                sheet.cell(row=2, column=col_idx, value=float(tensor.item()))
            elif tensor.dim() == 1:  # 1D tensor
                for row_idx, value in enumerate(tensor.tolist(), start=2):
                    sheet.cell(row=row_idx, column=col_idx, value=float(value))
            else:  # 2D+ tensor - flatten
                flat = tensor.flatten()
                for row_idx, value in enumerate(flat.tolist(), start=2):
                    sheet.cell(row=row_idx, column=col_idx, value=float(value))
        
        # Set column widths dynamically based on number of columns
        for col_idx in range(1, num_lists + 1):
            sheet.column_dimensions[chr(64 + col_idx)].width = 18
    
    wb.save(output_path)
    return output_path